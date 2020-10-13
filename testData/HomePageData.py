import openpyxl


class HomePageData:
    home_page_test_data = [{"name": "Chandler Bing", "email": "cbing@gmail.com", "gender": "Male"},
                           {"name": "Monica Geller", "email": "mgeller@gmail.com", "gender": "Female"}]

    @staticmethod
    def home_page_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\Users\\ratnakar_pc\PycharmProjects\PythonSeleniumFramework\ExcelFiles\data.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]